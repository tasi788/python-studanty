import openpyxl


def main():
    wb = openpyxl.load_workbook('names_and_savings.xlsx')
    ws = wb['工作表1']
    counter = 1
    for row in ws.iter_rows():
        numbers: int = row[2].value
        if numbers >= 20000:
            ws.cell(row=counter, column=4).value = 'VIP'
        counter += 1
    wb.save('names_and_savings_new.xlsx')


if __name__ == '__main__':
    main()



#print(results)

#print(ws.cell(row=1, column=1).value)
#print(ws.cell(row=1, column=2).value)
#print(ws.cell(row=1, column=3).value)
#print(type(ws.cell(row=1, column=1).value))
#print(type(ws.cell(row=1, column=2).value))
#print(type(ws.cell(row=1, column=3).value))
    


